import os
from openpyxl import Workbook, load_workbook
from blogs.models import BlogUser
from Blog.settings import MEDIA_ROOT

# 获得类的字段名称列表
def getClassFields():
    blogUserField = BlogUser.__dict__['__doc__']
    print("类字段：",blogUserField)
    blogUserField = BlogUser.__doc__
    print("类字段：",blogUserField)
    className = BlogUser.__name__
    classNameLength = len(className)
    # string = blogUserField[9:-1]
    string = blogUserField[classNameLength+1:-1]
    fieldList = string.split(", ")
    print("类名称：",className)
    print("类名称长度：",classNameLength)
    print("fields：",fieldList)
    # saveToExcel(fieldList)
    getFromExcel()
    return fieldList

# 将数据写入excel表
def saveToExcel(fieldList):
    # 创建一个工作簿
    wb = Workbook()
    # 创建一个sheet工作表
    ws = wb.create_sheet(title="TestExcel",index=0)
    ws = wb.active
    ws.append(fieldList)
    print(ws,wb.sheetnames,wb.sheetnames[0])
    filename = os.path.join(MEDIA_ROOT,"TestFile.xlsx")
    print("Excel文件名：",filename)
    wb.save(filename)

# 读出excel数据
def getFromExcel():
    filename = os.path.join(MEDIA_ROOT,"TestFile.xlsx")
    lwb = load_workbook(filename)
    ws = lwb.active
    # ws = lwb["TestExcel"]
    print("Excel表格信息：",ws,ws.title)
    rows = ws.rows
    # print("rows：",rows)
    rowsList = []
    for row in rows:
        # print("row：",row)
        cellsList = []
        for cell in row:
            value = cell.value
            # print("cell：",value)
            cellsList.append(value)
        rowsList.append(cellsList)
        # print("cellsList信息：",cellsList)
    # print("rowsList信息：",rowsList)
    rowsValueList = rowsList[1:]
    # print("rowsValueList信息：",rowsValueList)
    insertIntoTable(rowsValueList)

# excel表格数据写入数据库
def insertIntoTable(rowsValueList):
    failedCount = 0
    succeedCount = 0
    for row in rowsValueList:
        # print("行内容信息：",row)
        # print("行内容信息：",row[1])
        blogUserExist = BlogUser.objects.filter(username=row[1])
        # print("是否存在===>>：",blogUserExist)
        if blogUserExist.exists():
            # print(blogUserExist,"存在")
            failedCount += 1
        else:
            blogUser = BlogUser()
            # print(blogUserExist,"不存在")
            print(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11])
            blogUser.username=row[1]
            if row[2]==None:
                blogUser.password='666666'
            else:
                blogUser.password=row[2]
            blogUser.avatart=row[3]
            blogUser.birth=row[4]
            if row[5]==None:
                blogUser.sex=""
            else:
                blogUser.sex=row[5]
            if row[6]==None:
                blogUser.phone=""
            else:
                blogUser.phone=row[6]
            if row[7]==None:
                blogUser.email=""
            else:
                blogUser.email=row[7]
            if row[8]==None:
                blogUser.last_login='2000-01-01 00:00:00+00:00'
            else:
                blogUser.last_login=row[8]
            if row[9]==None:
                blogUser.is_staff=False
            else:
                blogUser.is_staff=row[9]
            if row[10]==None:
                blogUser.is_active=True
            else:
                blogUser.is_active=row[10]
            if row[11]==None:
                blogUser.date_joined='2000-01-01 00:00:00+00:00'
            else:
                blogUser.date_joined=row[11]
            blogUser.save()
            succeedCount += 1
    print("excel数据插入UserBlog表成功，",succeedCount,"条插入成功，",failedCount,"条插入失败")
